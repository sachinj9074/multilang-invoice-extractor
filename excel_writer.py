from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

COLUMNS = [
    "Source Filename",
    "Date Processed",
    "Invoice Number",
    "Invoice Date",
    "Due Date",
    "PO Number",
    "Detected Language",
    "Currency",
    "Vendor Name",
    "Vendor Address",
    "Vendor Tax ID",
    "Vendor Contact",
    "Buyer Name",
    "Buyer Address",
    "Buyer Tax ID",
    "Subtotal",
    "Tax Amount",
    "Tax Type",
    "Discount",
    "Total Amount",
    "Payment Terms",
    "Confidence",
    "Flags",
    "Needs Review",
]

FIELD_TO_COLUMN = {
    "invoice_number": "Invoice Number",
    "invoice_date": "Invoice Date",
    "due_date": "Due Date",
    "po_number": "PO Number",
    "detected_language": "Detected Language",
    "currency": "Currency",
    "vendor_name": "Vendor Name",
    "vendor_address": "Vendor Address",
    "vendor_tax_id": "Vendor Tax ID",
    "vendor_contact": "Vendor Contact",
    "buyer_name": "Buyer Name",
    "buyer_address": "Buyer Address",
    "buyer_tax_id": "Buyer Tax ID",
    "subtotal": "Subtotal",
    "tax_amount": "Tax Amount",
    "tax_type": "Tax Type",
    "discount": "Discount",
    "total_amount": "Total Amount",
    "payment_terms": "Payment Terms",
    "confidence": "Confidence",
}


def build_row(data, filename):
    row = {column: "" for column in COLUMNS}
    row["Source Filename"] = filename
    row["Date Processed"] = date.today().isoformat()

    if "error" in data:
        row["Flags"] = data["error"]
    else:
        for field, column in FIELD_TO_COLUMN.items():
            value = data.get(field)
            row[column] = value if value is not None else ""

        flags = data.get("flags") or []
        row["Flags"] = "; ".join(flags)

    row["Needs Review"] = data.get("needs_review", "")
    return row


def append_invoice_row(data, filename, excel_path="invoices.xlsx"):
    path = Path(excel_path)

    if path.exists():
        workbook = load_workbook(path)
        sheet = workbook.active

        existing_header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        if existing_header != COLUMNS:
            raise ValueError(
                f"{excel_path} has an outdated header row that doesn't match "
                f"the current column list. Refusing to append or modify it. "
                f"Existing header: {existing_header}. Expected: {COLUMNS}"
            )
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(COLUMNS)

    row = build_row(data, filename)
    sheet.append([row[column] for column in COLUMNS])
    workbook.save(path)
